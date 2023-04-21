import requests
from bs4 import BeautifulSoup
import openpyxl

# nomadcoder_List_of_course.py
# 노션 표에 타이틀과 서브 타이틀 네이밍을 쉽게 복붙하기 위해 만들었습니다.

wb = openpyxl.Workbook()
ws = wb.active

res = requests.get('https://nomadcoders.co/javascript-for-beginners/lobby')
soup = BeautifulSoup(res.content, 'html.parser')

course_title = soup.find("h3", {"class": "text-xl overflow-hidden"}).get_text()
ws['A1'] = course_title

titles = soup.find_all("span", {"class": "px-6 py-3 block min-w-full border-b border-gray-200 dark:border-gray-900 dark:bg-gray-900 dark:text-white bg-white text-left text-sm leading-4 font-medium text-gray-800 uppercase tracking-wider"})

for i in range(len(titles)):
    ws.cell(row=i+2, column=1, value=titles[i].get_text())

sub_titles1 = soup.find_all("span", {"class": "px-6 py-4 whitespace-nowrap text-sm leading-5 overflow-hidden font-medium flex items-center text-gray-900 dark:text-white"})

for i in range(len(sub_titles1)):
    ws.cell(row=i+2, column=2, value=sub_titles1[i].get_text())

sub_titles2 = soup.find_all("span", {"class": "px-6 py-4 whitespace-nowrap text-sm leading-5 overflow-hidden font-medium flex items-center text-gray-400"})

for i in range(len(sub_titles2)):
    ws.cell(row=i+2, column=3, value=sub_titles2[i].get_text())

title_list = []
for title in titles:
    title_list.append(title.get_text())
    print(title.get_text())
    ws.cell(row=len(title_list)+2, column=1, value=title.get_text())

wb.save("{}.xlsx".format(course_title))
